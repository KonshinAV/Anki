from deep_translator import GoogleTranslator
from gtts import gTTS
from g4f.client import Client
from pprint import pprint
import json
import uuid
import base64
import os
import requests
import openpyxl

class SimpleGoogleTranslate:
    def __init__(self, source='de', target='ru'):
        self.source = source
        self.target = target

    def translate(self, text):
        return GoogleTranslator(source=self.source, target=self.target).translate(text=text)

class Recorder:
    def make_record (text, save_to_path, lang = 'de'):
        tts = gTTS(text=text, lang=lang)
        tts.save(save_to_path)

class Anki:
    def __init__(self, ANKI_CONNECT_URL = 'http://localhost:8765', AUDIO_FOLDER = "tmp_audio", deck_name='Deutsche Lernen::Dev_deck', model_name_default='Basic (and reversed card)_main'):
        self.ANKI_CONNECT_URL = ANKI_CONNECT_URL
        self.deck_name = deck_name
        self.notes_ids = self.get_note_ids_from_deck()
        self.notes_info = self.get_notes_info()
        self.model_name_default = model_name_default

    def invoke(self, action, params=None):
        response = requests.post(self.ANKI_CONNECT_URL, json={
            'action': action,
            'version': 6,
            'params': params or {}
        }).json()
        if response.get("error") is not None:
            raise Exception(response["error"])
        return response["result"]

    def get_note_ids_from_deck(self):
        query = f'deck:"{self.deck_name}"'
        return self.invoke('findNotes', {'query': query})

    def get_notes_info(self,):
        return self.invoke('notesInfo', {'notes': self.notes_ids})

    def show_notes(self):
        # note_ids = self.get_note_ids_from_deck(self.deck_name)
        print(f"Найдено заметок в колоде '{self.deck_name}': {len(self.notes_ids)}")
        
        if not self.notes_ids:
            print("Нет заметок.")
            return
        
        notes = self.get_notes_info()
        for note in notes:
            pprint(note)

    def show_note_by_note_id(self, note_id):
        result = self.get_notes_info([note_id])
        if not result:
            print(f"Заметка с ID {note_id} не найдена.")
            return
        note = result[0]
        pprint(note)

    def update_note_field(self, note_id, field_name, new_value):
        # note_info = self.get_notes_info()
        if not self.notes_info:
            print(f"Заметка с ID {note_id} не найдена.")
            return

        # Проверка существования поля
        fields = self.notes_info[0]['fields']
        if field_name not in fields:
            print(f"Поле '{field_name}' не найдено в заметке.")
            return

        # Обновление поля
        self.invoke('updateNoteFields', {
            'note': {
                'id': note_id,
                'fields': {
                    field_name: new_value
                }
            }
        })
        print(f"Поле '{field_name}' обновлено на: {new_value}")

    def translate_base(self, de_field, ru_field, update_only_empty_values = True):
        notes_count = len(self.notes_info)
        for ind, note in enumerate(self.notes_info):
            note_id = note['noteId']
            de_value = note['fields'][de_field]['value']
            ru_value = note['fields'][ru_field]['value']
            print(f"{ind+1} von {notes_count}, {de_field}: {de_value}")
            if de_value != '':
                if update_only_empty_values:
                    if ru_value == '':
                        self.update_note_field(note_id=note_id,
                                            field_name=ru_field,
                                            new_value=translator.translate(text=de_value))
                    else:
                        print(f"Поле {ru_field} не пустое и будет пропущено. Текущее значение: {ru_value}")
                else:
                    self.update_note_field(note_id=note_id,
                                            field_name=ru_field,
                                            new_value=translator.translate(text=de_value))
            else:
                print(f"Field {de_field} is empty")

    def generate_and_insert_tts(self, notes, source_field, target_field):
        ''' Без API ключа не работает'''
        for note in notes:
            note_id = note['noteId']
            note_info = self.get_notes_info([note_id])
            if not note_info:
                print(f"Заметка с ID {note_id} не найдена.")
                return

            note = note_info[0]
            fields = note['fields']
            if source_field not in fields:
                print(f"Поле '{source_field}' не найдено.")
                return

            text = fields[source_field]['value']
            if not text.strip():
                print(f"Поле '{source_field}' пустое.")
                return

            # Генерация уникального имени файла
            uid = str(uuid.uuid4())  # Убираем дефисы для компактности
            filename = f"google-{uid}.mp3"

            print(f"Генерация озвучки для текста: {text}")
            print(f"Имя файла: {filename}")

            # Генерация аудио-файла
            self.invoke("ttsMake", {
                        "text": text,
                        "voice": "GoogleTranslate:de",
                        "cache": True,
                        "slow": False,
                        "useCacheOnly": False,
                        "filename": filename
                    })

            # Вставка [sound:filename] в нужное поле
            sound_tag = f"[sound:{filename}]"

            result = self.invoke("updateNoteFields", {
                "note": {
                    "id": note_id,
                    "fields": {
                        target_field: sound_tag
                    }
                }
            })

            # Добавление самого аудиофайла в заметку
            self.invoke("addNoteAudio", {
                "note": {
                    "id": note_id
                },
                "audio": {
                    "filename": filename,
                    "fields": [target_field],
                    "skipHash": "7e2c2f954ef6051373ba916f000168dc"
                }
            })

    def note_exists_in_deck(self, field_name, value, model_name=None):
        if model_name == None: model_name = self.model_name_default
        query = f'deck:"{self.deck_name}" note:"{model_name}" "{field_name}:{value}"'
        note_ids = self.invoke("findNotes", {"query": query})
        return len(note_ids) > 0

    def add_note_to_deck(self, fields, model_name=None, tags=None):
        if model_name == None: model_name = self.model_name_default
        note = {
            "deckName": self.deck_name,
            "modelName": model_name,
            "fields": fields,
            "options": {
                "allowDuplicate": False,
                "duplicateScope": "deck"
            },
            "tags": tags or [],
            "audio": [],
            "picture": [],
            "video": []
        }

        result = self.invoke("addNote", {"note": note})
        print(f"Добавлена заметка. ID: {result}")
        return result

    def add_audio_to_note(self, note_id, mp3_path, field_name):
        if not os.path.exists(mp3_path):
            raise FileNotFoundError(f"Файл не найден: {mp3_path}")

        filename = os.path.basename(mp3_path)
        # sound_tag = f"[sound:record-{filename}]"
        sound_tag = f"[sound:{filename}]"

        with open(mp3_path, "rb") as f:
            audio_data = f.read()
        b64_audio = base64.b64encode(audio_data).decode("utf-8")

        # Загружаем файл в медиа Anki
        self.invoke("storeMediaFile", {
            "filename": filename,
            "data": b64_audio
            })

        # Обновляем поле заметки
        self.invoke("updateNoteFields", {
            "note": {
                "id": note_id,
                "fields": {
                    field_name: sound_tag
                    }
            }
            })
        # Удаляем файл после добавления
        try:
            os.remove(mp3_path)
            # print(f"Файл '{mp3_path}' удалён.")
        except Exception as e:
            print(f"Не удалось удалить файл: {e}")

def read_xlsx_file(filename, sheet_name):
    # Загружаем файл
    workbook = openpyxl.load_workbook(filename)

    # Проверяем, существует ли лист с указанным именем
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Лист с именем '{sheet_name}' не найден в файле.")

    # Получаем нужный лист
    sheet = workbook[sheet_name]

    # Получаем заголовки из первой строки
    header = [cell for cell in next(sheet.iter_rows(values_only=True))]

    # Читаем остальные строки и формируем список словарей, пропуская пустые строки
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # Проверка на непустую строку
            row_dict = dict(zip(header, row))
            data.append(row_dict)

    return data

def add_notes_base_model(list_notes):
    for note in list_notes:
        if not anki.note_exists_in_deck(field_name="base_de", value=note['full_de']):
            for field in note:
                if note[field] == None:
                    note[field] = ''

            anki.add_note_to_deck(
                    fields={
                        "full_de": note['full_de'],
                        "base_de": note['base_de'],
                        "base_ru": note['base_ru'],
                        "artikel_de": note['artikel_de'],
                        "plural_de": note['plural_de'],
                        "notes": note['notes'],
                        "audio_text_de": note['audio_text_de'],
                        "s1_de": note['s1_de'],
                        "s1_ru": note['s1_ru'],
                        "s2_de": note['s2_de'],
                        "s2_ru": note['s2_ru'],
                        "s3_de": note['s3_de'],
                        "s3_ru": note['s3_ru'],
                        "s4_de": note['s4_de'],
                        "s4_ru": note['s4_ru'],
                        "s5_de": note['s5_de'],
                        "s5_ru": note['s5_ru'],
                        "s6_de": note['s6_de'],
                        "s6_ru": note['s6_ru'],
                        "s7_de": note['s7_de'],
                        "s7_ru": note['s7_ru'],
                        "s8_de": note['s8_de'],
                        "s8_ru": note['s8_ru'],
                        "s9_de": note['s9_de'],
                        "s9_ru": note['s9_ru'],
                        
                    },
                    tags=[]
                    )
        else:
            print(f"Запись {note["base_de"]} is already exists")
            pass

def gpt_text(content):
    client = Client()
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": str(content)}],
        web_search=False
    )
    return response.choices[0].message.content

def make_audio_records(source_field='s7_de', dest_field='s7_audio'):
    notes_ids = anki.notes_info
    for ind, note in enumerate(notes_ids):
        print (note['fields']['base_de']['value'])
        if not note['fields'][source_field]['value'] == '':
            noteId = note['noteId']
            audio_file_name = f"record-{uuid.uuid4()}.mp3"
            tts = gTTS(text=note['fields'][source_field]['value'], lang='de')
            tts.save(f"{audio_file_name}")
            anki.add_audio_to_note(note_id=noteId, mp3_path=f"{audio_file_name}", field_name=dest_field)
        else:
            print(f"Поле {source_field} для ноты {note['noteId']} пустое")

if __name__ == "__main__":
    translator = SimpleGoogleTranslate()
    # deck_name = "Deutsche Lernen::B1_Wortliste_lernen"
    deck_name = "Deutsche Lernen::Wortschatz"
    # deck_name = "Deutsche Lernen::Dev_deck"
    # model_name = "Basic (and reversed card)_main"
    anki = Anki(deck_name=deck_name)
    xlsx_file_content = read_xlsx_file(filename="for_import.xlsx", sheet_name="Sheet2")
    # pprint(xlsx_file_content)
    
    # add_notes_base_model(xlsx_file_content)
    
    make_audio_records(source_field='s9_de',dest_field='s9_audio')

